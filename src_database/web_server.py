import os
import sys
import json
import glob
import datetime
import urllib.parse
import subprocess
import threading
import xml.etree.ElementTree as ET
import openpyxl
from http.server import HTTPServer, SimpleHTTPRequestHandler

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WEB_UI_DIR = os.path.join(BASE_DIR, 'src_database', 'web_ui')
TARGET_LINKS_FILE = os.path.join(BASE_DIR, 'src_database', 'target_links.txt')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
LOGS_DIR = os.path.join(BASE_DIR, 'logs')
REF_DIR = os.path.join(BASE_DIR, 'reference_files')
VD_POINT_LIST_FILE = os.path.join(REF_DIR, 'vd_point_list.xml')
if not os.path.exists(VD_POINT_LIST_FILE):
    VD_POINT_LIST_FILE = os.path.join(BASE_DIR, 'vd_point_list.xml')
IC_CSV_FILE   = os.path.join(REF_DIR, 'Freeway_Interchanges_Full.csv')
EXPR_CSV_FILE = os.path.join(REF_DIR, 'Expressway_Interchanges.csv')

# ── 路段類型分類（依路線名稱字串動態判斷）──────────────────────────────────────────
def categorize_xml_road(road_name):
    """將 XML RoadName 分類為 國道主線 / 國道支線 / 快速道路 / 其他."""
    if '快速公路' in road_name:
        return '快速道路'
    if '高架' in road_name or (road_name.startswith('國道') and ('甲' in road_name or '乙' in road_name)):
        return '國道支線'
    if road_name.startswith('國道'):
        return '國道主線'
    return '其他'

# XML RoadName → CSV 路線名稱對照
XML_TO_CSV_ROAD = {
    '國道1號':  '國道1號',
    '國1高架':  '汐五及五楊高架',  # XML 中的 國1高架 對應 CSV 汐五及五楊高架
    '台1甲':   '汐五及五楊高架',
    '國道2號':  '國道2號',
    '國道2甲':  '國道2甲',
    '國道3號':  '國道3號',
    '國道3甲':  '國道3甲',
    '國道4號':  '國道4號',
    '國道5號':  '國道5號',
    '國道6號':  '國道6號',
    '國道8號':  '國道8號',
    '國道10號': '國道10號',
    # 快速公路對應 Expressway_Interchanges.csv 中的 路線名稱
    '快速公路62號': '快速公路62號',
    '快速公路64號': '快速公路64號',
    '快速公路72號': '快速公路72號',
    '快速公路74甲': '快速公路74甲',
    '快速公路74號': '快速公路74號',
    '快速公路76號': '快速公路76號',
    '快速公路78號': '快速公路78號',
    '快速公路82號': '快速公路82號',
    '快速公路84號': '快速公路84號',
    '快速公路86號': '快速公路86號',
    '快速公路88號': '快速公路88號',
}

# Global State for Analysis Run
state_lock = threading.Lock()
current_run = {
    "status": "idle",
    "progress": 0,
    "message": "系統已就緒",
    "logs": []
}

def run_analysis_thread(date_str, mode, custom_hours):
    global current_run
    with state_lock:
        current_run["status"] = "running"
        current_run["progress"] = 0
        current_run["message"] = f"開始分析 ({date_str}, 模式: {mode})..."
        current_run["logs"] = [f"=== 開始執行交通量分析 [日期: {date_str}, 模式: {mode}] ==="]

    cmd = [sys.executable, os.path.join(BASE_DIR, 'src_database', 'generate_traffic_report.py'), date_str, mode, custom_hours]
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', errors='replace', bufsize=1)

    for line in iter(proc.stdout.readline, ''):
        line_str = line.strip()
        if not line_str:
            continue
        
        with state_lock:
            current_run["logs"].append(line_str)
            if line_str.startswith("PROGRESS:"):
                try:
                    parts = line_str.replace("PROGRESS:", "").split("|")
                    pct = int(parts[0].replace("%", "").strip())
                    msg = parts[1].strip() if len(parts) > 1 else ""
                    current_run["progress"] = pct
                    current_run["message"] = msg
                except Exception:
                    pass

    proc.stdout.close()
    return_code = proc.wait()

    with state_lock:
        if return_code == 0:
            current_run["status"] = "completed"
            current_run["progress"] = 100
            current_run["message"] = "分析完成！報表已產出。"
            current_run["logs"].append("=== 轉檔與分析順利完成 ===")
        else:
            current_run["status"] = "error"
            current_run["message"] = "執行過程發生錯誤！"
            current_run["logs"].append("=== 執行失敗，請檢查 Log ===")

def parse_latest_excel_results(requested_date=None):
    if not os.path.exists(OUTPUT_DIR):
        return None
        
    target_file = None
    if requested_date:
        candidate = os.path.join(OUTPUT_DIR, f"VD_traffic_report_{requested_date}.xlsx")
        if os.path.exists(candidate):
            target_file = candidate
            
    if not target_file:
        files = sorted(glob.glob(os.path.join(OUTPUT_DIR, "*.xlsx")), key=os.path.getmtime, reverse=True)
        if not files:
            return None
        target_file = files[0]

    filename = os.path.basename(target_file)
    
    try:
        wb = openpyxl.load_workbook(target_file, data_only=True)
        results = []
        highways = set()
        
        # 1. Sheet: 國道主線
        if '國道主線' in wb.sheetnames:
            ws = wb['國道主線']
            for r in range(3, ws.max_row + 1):
                road_name = str(ws.cell(row=r, column=1).value or '').strip()
                segment = str(ws.cell(row=r, column=2).value or '').strip()
                direction = str(ws.cell(row=r, column=3).value or '').strip()
                cap = float(ws.cell(row=r, column=4).value or 0)
                limit = float(ws.cell(row=r, column=5).value or 0)
                
                m_pcu = float(ws.cell(row=r, column=6).value or 0)
                m_vc = float(ws.cell(row=r, column=7).value or 0)
                m_spd = float(ws.cell(row=r, column=8).value or 0)
                m_los = str(ws.cell(row=r, column=9).value or '').strip()
                
                e_pcu = float(ws.cell(row=r, column=10).value or 0)
                e_vc = float(ws.cell(row=r, column=11).value or 0)
                e_spd = float(ws.cell(row=r, column=12).value or 0)
                e_los = str(ws.cell(row=r, column=13).value or '').strip()
                
                if road_name and segment:
                    highways.add(road_name)
                    results.append({
                        "road_name": road_name,
                        "segment": segment,
                        "direction": direction,
                        "type": "主線",
                        "capacity": cap,
                        "speed_limit": limit,
                        "m_pcu": m_pcu, "m_vc": m_vc, "m_speed": m_spd, "m_los": m_los,
                        "e_pcu": e_pcu, "e_vc": e_vc, "e_speed": e_spd, "e_los": e_los
                    })
                    
        # 2. Sheet: 國道匝道
        if '國道匝道' in wb.sheetnames:
            ws = wb['國道匝道']
            for r in range(3, ws.max_row + 1):
                road_name = str(ws.cell(row=r, column=1).value or '').strip()
                ic_name = str(ws.cell(row=r, column=2).value or '').strip()
                direction = str(ws.cell(row=r, column=3).value or '').strip()
                in_out = str(ws.cell(row=r, column=4).value or '').strip()
                dest = str(ws.cell(row=r, column=5).value or '').strip()
                cap = float(ws.cell(row=r, column=6).value or 0)
                limit = float(ws.cell(row=r, column=7).value or 0)
                
                m_pcu = float(ws.cell(row=r, column=8).value or 0)
                m_vc = float(ws.cell(row=r, column=9).value or 0)
                m_spd = float(ws.cell(row=r, column=10).value or 0)
                m_los = str(ws.cell(row=r, column=11).value or '').strip()
                
                e_pcu = float(ws.cell(row=r, column=12).value or 0)
                e_vc = float(ws.cell(row=r, column=13).value or 0)
                e_spd = float(ws.cell(row=r, column=14).value or 0)
                e_los = str(ws.cell(row=r, column=15).value or '').strip()
                
                if road_name and ic_name:
                    highways.add(road_name)
                    results.append({
                        "road_name": road_name,
                        "segment": f"{ic_name} ({in_out}-{dest})",
                        "direction": direction,
                        "type": "匝道",
                        "capacity": cap,
                        "speed_limit": limit,
                        "m_pcu": m_pcu, "m_vc": m_vc, "m_speed": m_spd, "m_los": m_los,
                        "e_pcu": e_pcu, "e_vc": e_vc, "e_speed": e_spd, "e_los": e_los
                    })
                    
        total_links = len(results)
        mainline_count = len([x for x in results if x['type'] == '主線'])
        ramp_count = len([x for x in results if x['type'] == '匝道'])
        
        max_vc_item = max(results, key=lambda x: max(x['m_vc'], x['e_vc'])) if results else None
        max_vc = max(max_vc_item['m_vc'], max_vc_item['e_vc']) if max_vc_item else 0
        max_vc_seg = f"{max_vc_item['segment']} ({max_vc_item['direction']})" if max_vc_item else ''
        
        worst_los_item = max(results, key=lambda x: max(x['m_los'], x['e_los'])) if results else None
        worst_los = max(worst_los_item['m_los'], worst_los_item['e_los']) if worst_los_item else 'A1'
        
        return {
            "report_name": filename,
            "total_links": total_links,
            "mainline_count": mainline_count,
            "ramp_count": ramp_count,
            "max_vc": round(max_vc, 2),
            "max_vc_seg": max_vc_seg,
            "worst_los": worst_los,
            "highways": sorted(list(highways)),
            "data": results
        }
    except Exception as e:
        print(f"Error parsing excel results: {e}")
        return None

import csv
import re

_ic_cache = None
_vd_xml_cache = None

def load_ic_data():
    """Load IC data from both Freeway and Expressway CSVs. Key = road name (exact 路線 value)."""
    global _ic_cache
    if _ic_cache is not None:
        return _ic_cache
    _ic_cache = {}  # road_name -> [{name, km}]

    def _load_csv(filepath, encoding='utf-8-sig'):
        if not os.path.exists(filepath):
            return
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                reader = csv.DictReader(f)
                cols = reader.fieldnames or []
                # Detect column names flexibly
                col_road = next((c for c in cols if '路線' in c or '路線' in c), cols[0] if cols else '')
                col_name = next((c for c in cols if '設施' in c or '名稱' in c), cols[1] if len(cols)>1 else '')
                col_km   = next((c for c in cols if '里程' in c), cols[2] if len(cols)>2 else '')
                for row in reader:
                    road = row.get(col_road, '').strip()
                    name = row.get(col_name, '').strip()
                    km_raw = row.get(col_km, '').strip()
                    if not road or not name:
                        continue
                    if name in ('A','B','C','D') or '次出口' in name or '休息站' in name or '服務區' in name:
                        continue
                    if name in ('設施名稱', '交流道數量合計', '路段數量合計'):
                        continue
                    km_str = re.sub(r'[^0-9.]', '', km_raw)
                    if not km_str:
                        continue
                    try:
                        km = float(km_str)
                    except ValueError:
                        continue
                    if road not in _ic_cache:
                        _ic_cache[road] = []
                    _ic_cache[road].append({'name': name, 'km': km})
        except Exception as e:
            print(f'IC CSV load error ({filepath}): {e}')

    _load_csv(IC_CSV_FILE)
    _load_csv(EXPR_CSV_FILE)
    return _ic_cache


def load_vd_xml_cache():
    """Parse vd_point_list.xml once and cache all links."""
    global _vd_xml_cache
    if _vd_xml_cache is not None:
        return _vd_xml_cache
    _vd_xml_cache = {}  # link_id -> {vd_id, road_name, direction, km, lanes, type}
    if not os.path.exists(VD_POINT_LIST_FILE):
        return _vd_xml_cache
    try:
        tree = ET.parse(VD_POINT_LIST_FILE)
        root = tree.getroot()
        ns = '{http://traffic.transportdata.tw/standard/traffic/schema/}'
        dir_text_map = {'E': '往東', 'W': '往西', 'N': '北上', 'S': '南下'}
        for vd in root.iter(f'{ns}VD'):
            vd_id = vd.findtext(f'{ns}VDID') or ''
            road_name = vd.findtext(f'{ns}RoadName') or ''
            loc_mile = vd.findtext(f'{ns}LocationMile') or ''
            # Parse km from LocationMile like '13K+550' -> 13.55
            m = re.match(r'(\d+)K\+(\d+)', loc_mile)
            km = (float(m.group(1)) + float(m.group(2)) / 1000) if m else 0.0
            for dlink in vd.iter(f'{ns}DetectionLink'):
                link_id = dlink.findtext(f'{ns}LinkID') or ''
                road_dir = dlink.findtext(f'{ns}RoadDirection') or ''
                lane_num = dlink.findtext(f'{ns}LaneNum') or '3'
                if not link_id:
                    continue
                feature_code = link_id[6] if len(link_id) >= 7 else '0'
                link_type = '主線' if feature_code == '0' else '匝道'
                dir_text = dir_text_map.get(road_dir, road_dir)
                _vd_xml_cache[link_id] = {
                    'vd_id': vd_id,
                    'road_name': road_name,
                    'direction': dir_text,
                    'road_dir': road_dir,
                    'km': km,
                    'lanes': lane_num,
                    'type': link_type
                }
    except Exception as e:
        print(f'VD XML cache error: {e}')
    return _vd_xml_cache


def get_browse_roads():
    """Return structured type → road → IC list for the browser UI."""
    ic_data = load_ic_data()
    vd_cache = load_vd_xml_cache()

    # Collect unique XML road names
    xml_roads = sorted(set(v['road_name'] for v in vd_cache.values() if v['road_name']))

    # Assign each XML road to a type category
    type_roads = {'國道主線': [], '國道支線': [], '快速道路': [], '其他': []}
    for xml_road in xml_roads:
        cat = categorize_xml_road(xml_road)
        type_roads[cat].append(xml_road)

    # Build IC list for each road
    result = {}
    for cat, roads in type_roads.items():
        if not roads:
            continue
        result[cat] = {}
        for xml_road in roads:
            csv_name = XML_TO_CSV_ROAD.get(xml_road, xml_road)
            ics = ic_data.get(csv_name, [])
            result[cat][xml_road] = sorted(ics, key=lambda x: x['km'])
    return result


def get_browse_links(road, link_type, direction, km_from, km_to):
    """Return filtered link list from XML cache."""
    vd_cache = load_vd_xml_cache()
    seen = set()
    links = []
    for link_id, info in sorted(vd_cache.items(), key=lambda x: x[1]['km']):
        if road and info['road_name'] != road:
            continue
        if link_type and link_type != '全部' and info['type'] != link_type:
            continue
        if direction and direction != '全部' and info['road_dir'] != direction:
            continue
        if km_from is not None and info['km'] < km_from:
            continue
        if km_to is not None and info['km'] > km_to:
            continue
        if link_id in seen:
            continue
        seen.add(link_id)
        links.append({
            'link_id': link_id,
            'vd_id': info['vd_id'],
            'road_name': info['road_name'],
            'direction': info['direction'],
            'type': info['type'],
            'lanes': info['lanes'],
            'km': round(info['km'], 3)
        })
    return links


def get_link_metadata_details():
    links = []
    if os.path.exists(TARGET_LINKS_FILE):
        with open(TARGET_LINKS_FILE, 'r', encoding='utf-8') as f:
            links = [line.strip() for line in f if line.strip()]
            
    xml_map = {}
    if os.path.exists(VD_POINT_LIST_FILE):
        try:
            tree = ET.parse(VD_POINT_LIST_FILE)
            root = tree.getroot()
            ns = '{http://traffic.transportdata.tw/standard/traffic/schema/}'
            for vd in root.iter(f'{ns}VD'):
                vd_id = vd.findtext(f'{ns}VDID')
                road_name = vd.findtext(f'{ns}RoadName') or '國道2號'
                for dlink in vd.iter(f'{ns}DetectionLink'):
                    link_id = dlink.findtext(f'{ns}LinkID')
                    road_dir = dlink.findtext(f'{ns}RoadDirection')
                    lane_num = dlink.findtext(f'{ns}LaneNum') or '3'
                    if link_id:
                        dir_text = '往東' if road_dir == 'E' else ('往西' if road_dir == 'W' else ('北上' if road_dir == 'N' else '南下'))
                        xml_map[link_id] = {
                            "vd_id": vd_id,
                            "road_name": road_name,
                            "direction": dir_text,
                            "lanes": lane_num
                        }
        except Exception:
            pass
            
    details = []
    for lid in links:
        feature_code = lid[6] if len(lid) >= 7 else '0'
        type_text = '主線' if feature_code == '0' else '匝道'
        info = xml_map.get(lid, {
            "vd_id": "VD-N2-E-LOOP" if lid.startswith("00002010") else ("VD-N2-W-LOOP" if lid.startswith("00002011") else "靜態資料庫未收錄"),
            "road_name": "國道2號",
            "direction": "往東" if lid[7] == '0' else "往西",
            "lanes": "3"
        })
        details.append({
            "link_id": lid,
            "vd_id": info["vd_id"],
            "road_name": info["road_name"],
            "direction": info["direction"],
            "type": type_text,
            "lanes": info["lanes"]
        })
    return details

class VDRequestHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=WEB_UI_DIR, **kwargs)

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        query = urllib.parse.parse_qs(parsed.query)

        if path == '/api/latest_results':
            req_date = query.get('date', [''])[0]
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            res = parse_latest_excel_results(req_date)
            self.wfile.write(json.dumps(res or {}, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/link_metadata':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            details = get_link_metadata_details()
            self.wfile.write(json.dumps({"metadata": details}, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/progress':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            with state_lock:
                self.wfile.write(json.dumps(current_run, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/links':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            links = []
            if os.path.exists(TARGET_LINKS_FILE):
                with open(TARGET_LINKS_FILE, 'r', encoding='utf-8') as f:
                    links = [line.strip() for line in f if line.strip()]
            self.wfile.write(json.dumps({"links": links}, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/reports':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            reports = []
            if os.path.exists(OUTPUT_DIR):
                files = sorted(glob.glob(os.path.join(OUTPUT_DIR, "*.xlsx")), key=os.path.getmtime, reverse=True)
                for f in files:
                    reports.append({
                        "name": os.path.basename(f),
                        "size": f"{os.path.getsize(f) / 1024:.1f} KB",
                        "mtime": datetime.datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M:%S")
                    })
            self.wfile.write(json.dumps({"reports": reports}, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/download':
            file_name = query.get('file', [''])[0]
            file_path = os.path.join(OUTPUT_DIR, os.path.basename(file_name))
            if os.path.exists(file_path):
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(file_name)}')
                self.send_header('Content-Length', str(os.path.getsize(file_path)))
                self.end_headers()
                with open(file_path, 'rb') as f:
                    self.wfile.write(f.read())
            else:
                self.send_error(404, "File Not Found")
        elif path == '/api/logs':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            logs = []
            if os.path.exists(LOGS_DIR):
                files = sorted(glob.glob(os.path.join(LOGS_DIR, "*.log")), key=os.path.getmtime, reverse=True)
                for f in files:
                    logs.append({
                        "name": os.path.basename(f),
                        "size": f"{os.path.getsize(f) / 1024:.1f} KB",
                        "mtime": datetime.datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M:%S")
                    })
            self.wfile.write(json.dumps({"logs": logs}, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/log_content':
            file_name = query.get('file', [''])[0]
            file_path = os.path.join(LOGS_DIR, os.path.basename(file_name))
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            content = ""
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
            self.wfile.write(json.dumps({"content": content}, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/references':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            refs = []
            if os.path.exists(REF_DIR):
                for fname in sorted(os.listdir(REF_DIR)):
                    fpath = os.path.join(REF_DIR, fname)
                    if os.path.isfile(fpath):
                        refs.append({
                            "name": fname,
                            "size": f"{os.path.getsize(fpath) / 1024:.1f} KB",
                            "ext": os.path.splitext(fname)[1].lower()
                        })
            self.wfile.write(json.dumps({"references": refs}, ensure_ascii=False).encode('utf-8'))
        elif path == '/api/browse_roads':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(get_browse_roads(), ensure_ascii=False).encode('utf-8'))
        elif path == '/api/browse_links':
            road = urllib.parse.unquote(query.get('road', [''])[0])
            link_type = urllib.parse.unquote(query.get('type', ['全部'])[0])
            direction = urllib.parse.unquote(query.get('dir', ['全部'])[0])
            km_from_s = query.get('km_from', [''])[0]
            km_to_s = query.get('km_to', [''])[0]
            km_from = float(km_from_s) if km_from_s else None
            km_to = float(km_to_s) if km_to_s else None
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            links = get_browse_links(road, link_type, direction, km_from, km_to)
            self.wfile.write(json.dumps({'links': links}, ensure_ascii=False).encode('utf-8'))
        else:
            super().do_GET()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)

        if path == '/api/run':
            data = json.loads(post_data.decode('utf-8'))
            date_str = data.get('date', '20260716')
            mode = data.get('mode', 'peak')
            custom_hours = data.get('custom_hours', '')

            with state_lock:
                if current_run["status"] == "running":
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json; charset=utf-8')
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "目前已有任務正在執行中！"}, ensure_ascii=False).encode('utf-8'))
                    return

            t = threading.Thread(target=run_analysis_thread, args=(date_str, mode, custom_hours), daemon=True)
            t.start()

            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps({"status": "started"}, ensure_ascii=False).encode('utf-8'))

        elif path == '/api/save_links':
            data = json.loads(post_data.decode('utf-8'))
            links = data.get('links', [])
            with open(TARGET_LINKS_FILE, 'w', encoding='utf-8') as f:
                for lid in links:
                    if lid.strip():
                        f.write(lid.strip() + '\n')
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps({"status": "success"}, ensure_ascii=False).encode('utf-8'))

def run_server(port=8000):
    os.makedirs(WEB_UI_DIR, exist_ok=True)
    server_address = ('', port)
    httpd = HTTPServer(server_address, VDRequestHandler)
    print(f"==================================================")
    print(f"VD Traffic Report Web UI Server Started!")
    print(f"Please open browser at: http://localhost:{port}")
    print(f"Press Ctrl+C to stop server")
    print(f"==================================================")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")

if __name__ == '__main__':
    port_input = int(sys.argv[1]) if len(sys.argv) > 1 and sys.argv[1].isdigit() else 8000
    run_server(port_input)
