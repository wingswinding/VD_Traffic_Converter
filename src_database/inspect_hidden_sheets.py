import openpyxl
import os

def inspect_visibility(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=False)
    
    output = []
    output.append(f"Workbook: {file_path}")
    output.append(f"All Sheet Names: {wb.sheetnames}")
    
    for name in wb.sheetnames:
        sheet = wb[name]
        output.append(f"Sheet Name: '{name}', State: '{sheet.sheet_state}'")
        
    wb.close()
    
    out_path = r"C:\Users\Owner\.gemini\antigravity\brain\5da4f6b9-54ae-48d5-bc02-be8a5a3438dd\scratch\hidden_sheets_status.txt"
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(output))
    print("Done. Written to", out_path)

if __name__ == '__main__':
    inspect_visibility('國道LOS.xlsx')
