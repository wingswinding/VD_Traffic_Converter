import os
import sys
import gzip
import glob
import shutil
import urllib3
import requests
import datetime
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TARGET_LINKS_FILE = os.path.join(BASE_DIR, 'src_database', 'target_links.txt')
DOWNLOAD_DIR = os.path.join(BASE_DIR, 'src_database', 'downloads')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
BACKUP_DIR = os.path.join(BASE_DIR, 'backup')
LOGS_DIR = os.path.join(BASE_DIR, 'logs')

VD_POINT_LIST_FILE = os.path.join(BASE_DIR, 'reference_files', 'vd_point_list.xml')
if not os.path.exists(VD_POINT_LIST_FILE):
    VD_POINT_LIST_FILE = os.path.join(BASE_DIR, 'vd_point_list.xml')

os.makedirs(DOWNLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)

class DualLogger:
    def __init__(self, log_filepath):
        self.terminal = sys.stdout
        self.log = open(log_filepath, "w", encoding="utf-8")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
        self.log.flush()

    def flush(self):
        self.terminal.flush()
        self.log.flush()

def report_progress(percent, msg):
    print(f"PROGRESS: {percent}% | {msg}")
    sys.stdout.flush()

def backup_file_if_exists(file_path):
    if os.path.exists(file_path):
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{name}_{timestamp}{ext}"
        backup_path = os.path.join(BACKUP_DIR, backup_filename)
        shutil.copy2(file_path, backup_path)
        print(f"[Backup] Existing file backed up to: {backup_path}")
        return backup_path
    return None

def get_free_speed(speed_limit):
    if speed_limit >= 110:
        return 115
    elif speed_limit >= 100:
        return 105
    elif speed_limit >= 90:
        return 100
    else:
        return 90

def get_mainline_capacity(speed_limit, lanes, open_shoulder):
    vf = get_free_speed(speed_limit)
    has_shoulder = (str(open_shoulder).upper() == 'Y')
    effective_lanes = lanes + (1 if has_shoulder else 0)
    
    if not has_shoulder:
        if effective_lanes == 4:
            qmax = 1950 if vf == 115 else (1900 if vf == 110 else (1850 if vf == 105 else 1800))
        elif effective_lanes == 3:
            qmax = 2000 if vf == 115 else (1950 if vf == 110 else (1900 if vf == 105 else 1850))
        elif effective_lanes == 2:
            qmax = 2050 if vf == 115 else (2000 if vf == 110 else (1950 if vf == 105 else 1900))
        else:
            qmax = 1850
    else:
        if lanes == 3:
            qmax = 1760 if vf == 115 else (1725 if vf == 110 else (1690 if vf == 105 else 1650))
        elif lanes == 2:
            qmax = 1730 if vf == 115 else (1700 if vf == 110 else (1670 if vf == 105 else 1630))
        else:
            qmax = 1690
            
    return qmax * effective_lanes

def get_ramp_capacity(ramp_type, lanes):
    if '出口' in ramp_type or 'OFF' in ramp_type.upper() or 'EXIT' in ramp_type.upper():
        return 3800 if lanes >= 2 else 1900
    else:
        return 3000 if lanes >= 2 else 1800

def calculate_los(vc_ratio, speed, speed_limit):
    if vc_ratio <= 0.25:
        letter = 'A'
    elif vc_ratio <= 0.50:
        letter = 'B'
    elif vc_ratio <= 0.80:
        letter = 'C'
    elif vc_ratio <= 0.90:
        letter = 'D'
    elif vc_ratio <= 1.00:
        letter = 'E'
    else:
        letter = 'F'
        
    speed_ratio = speed / speed_limit if speed_limit > 0 else 0
    if speed_ratio >= 0.90:
        num = '1'
    elif speed_ratio >= 0.80:
        num = '2'
    elif speed_ratio >= 0.60:
        num = '3'
    elif speed_ratio >= 0.40:
        num = '4'
    elif speed_ratio >= 0.20:
        num = '5'
    else:
        num = '6'
        
    return f"{letter}{num}"

def download_vd_data(date_str, hours):
    date_dir = os.path.join(DOWNLOAD_DIR, date_str)
    os.makedirs(date_dir, exist_ok=True)
    
    total_files = len(hours) * 60
    downloaded = 0
    
    for h in hours:
        for m in range(0, 60):
            time_str = f"{h:02d}{m:02d}"
            filename = f"VDLive_{time_str}.xml.gz"
            local_path = os.path.join(date_dir, filename)
            
            if not os.path.exists(local_path) or os.path.getsize(local_path) == 0:
                url = f"https://tisvcloud.freeway.gov.tw/history/motc20/VD/{date_str}/{filename}"
                try:
                    resp = requests.get(url, timeout=10, verify=False)
                    if resp.status_code == 200:
                        with open(local_path, 'wb') as f:
                            f.write(resp.content)
                except Exception as e:
                    pass
            
            downloaded += 1
            if downloaded % 20 == 0 or downloaded == total_files:
                prog_pct = 10 + int((downloaded / max(total_files, 1)) * 35) # 10% to 45%
                report_progress(prog_pct, f"Downloading XML files ({downloaded}/{total_files})...")
                
    return date_dir

def parse_custom_blocks(custom_hours_str):
    """Parse custom hours string into consecutive hour blocks."""
    blocks = []
    parts = [p.strip() for p in custom_hours_str.split(',') if p.strip()]
    for part in parts:
        if '-' in part:
            try:
                s_h, e_h = map(int, part.split('-'))
                hours = [h for h in range(s_h, e_h) if 0 <= h < 24]
                if hours:
                    blocks.append(hours)
            except ValueError:
                pass
        elif part.isdigit():
            h = int(part)
            if 0 <= h < 24:
                if blocks and blocks[-1][-1] == h - 1:
                    blocks[-1].append(h)
                else:
                    blocks.append([h])
    return blocks

def process_vd_data(date_dir, target_links, hours_config, is_weekend=False, mode='peak', custom_hours_str='', pce_s=1.0, pce_l=1.8, pce_t=2.5):
    hours_list, period_map = hours_config
    
    hourly_data = {h: {lid: {
        'S': 0, 'S_speed_vol': 0,
        'L': 0, 'L_speed_vol': 0,
        'T': 0, 'T_speed_vol': 0,
        'speed_vol': 0, 'vol': 0
    } for lid in target_links} for h in hours_list}

    # Store minute-by-minute data for 15-minute sliding 1-hour window peak calculation
    minute_data = {}
    
    link_vdid_map = {}
    
    gz_files = glob.glob(os.path.join(date_dir, "*.xml.gz"))
    total_gz = len(gz_files)
    print(f"Processing {total_gz} XML.GZ files...")
    report_progress(50, f"Parsing {total_gz} XML data files...")
    
    ns = '{http://traffic.transportdata.tw/standard/traffic/schema/}'
    
    proc_count = 0
    for gz_file in gz_files:
        basename = os.path.basename(gz_file)
        time_part = basename.replace("VDLive_", "").replace(".xml.gz", "")
        if len(time_part) != 4 or not time_part.isdigit():
            continue
        hour = int(time_part[:2])
        minute = int(time_part[2:])
        m_of_day = hour * 60 + minute
        
        h_key = period_map.get(hour)
        
        try:
            with gzip.open(gz_file, 'rb') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                
                for vd_live in root.iter(f'{ns}VDLive'):
                    vd_id = vd_live.findtext(f'{ns}VDID')
                    for dlink in vd_live.iter(f'{ns}LinkFlow'):
                        link_id = dlink.findtext(f'{ns}LinkID')
                        if link_id in target_links:
                            if link_id not in link_vdid_map and vd_id:
                                link_vdid_map[link_id] = vd_id
                                
                            if m_of_day not in minute_data:
                                minute_data[m_of_day] = {}
                            if link_id not in minute_data[m_of_day]:
                                minute_data[m_of_day][link_id] = {'S': 0, 'L': 0, 'T': 0, 'speed_vol': 0, 'vol': 0}

                            for lane in dlink.iter(f'{ns}Lane'):
                                speed = float(lane.findtext(f'{ns}Speed') or 0)
                                for vehicle in lane.iter(f'{ns}Vehicle'):
                                    vtype = vehicle.findtext(f'{ns}VehicleType')
                                    vol = float(vehicle.findtext(f'{ns}Volume') or 0)
                                    veh_speed = float(vehicle.findtext(f'{ns}Speed') or speed)
                                    
                                    if vtype == 'S':
                                        if h_key:
                                            hourly_data[h_key][link_id]['S'] += vol
                                            hourly_data[h_key][link_id]['S_speed_vol'] += veh_speed * vol
                                        minute_data[m_of_day][link_id]['S'] += vol
                                    elif vtype == 'L':
                                        if h_key:
                                            hourly_data[h_key][link_id]['L'] += vol
                                            hourly_data[h_key][link_id]['L_speed_vol'] += veh_speed * vol
                                        minute_data[m_of_day][link_id]['L'] += vol
                                    elif vtype == 'T':
                                        if h_key:
                                            hourly_data[h_key][link_id]['T'] += vol
                                            hourly_data[h_key][link_id]['T_speed_vol'] += veh_speed * vol
                                        minute_data[m_of_day][link_id]['T'] += vol
                                        
                                    if h_key:
                                        hourly_data[h_key][link_id]['speed_vol'] += veh_speed * vol
                                        hourly_data[h_key][link_id]['vol'] += vol
                                    minute_data[m_of_day][link_id]['speed_vol'] += veh_speed * vol
                                    minute_data[m_of_day][link_id]['vol'] += vol
        except Exception:
            pass
            
        proc_count += 1
        if proc_count % 30 == 0 or proc_count == total_gz:
            prog_pct = 50 + int((proc_count / max(total_gz, 1)) * 30) # 50% to 80%
            report_progress(prog_pct, f"Parsing XML records ({proc_count}/{total_gz})...")

    final_peak_metrics = {lid: {'morning': {'pcu': 0, 'speed': 0}, 'evening': {'pcu': 0, 'speed': 0}} for lid in target_links}
    
    use_sliding = False
    m_starts = []
    e_starts = []

    if mode == 'peak':
        use_sliding = True
        if is_weekend:
            m_starts = [600, 615, 630, 645, 660]    # 10:00 ~ 12:00
            e_starts = [960, 975, 990, 1005, 1020]   # 16:00 ~ 18:00
        else:
            m_starts = [420, 435, 450, 465, 480]     # 07:00 ~ 09:00
            e_starts = [1020, 1035, 1050, 1065, 1080] # 17:00 ~ 19:00
    elif mode == 'custom' and custom_hours_str:
        blocks = parse_custom_blocks(custom_hours_str)
        # Rule: 4小時內使用15分鐘間隔最大量；4小時以上使用獨立小時
        if blocks and all(len(b) <= 4 for b in blocks):
            use_sliding = True
            b1 = blocks[0]
            start_m1 = b1[0] * 60
            end_m1 = (b1[-1] + 1) * 60
            m_starts = list(range(start_m1, max(start_m1 + 1, end_m1 - 60 + 1), 15))

            if len(blocks) > 1:
                b2 = blocks[1]
                start_m2 = b2[0] * 60
                end_m2 = (b2[-1] + 1) * 60
                e_starts = list(range(start_m2, max(start_m2 + 1, end_m2 - 60 + 1), 15))

    if use_sliding and m_starts:
        for lid in target_links:
            # Morning Peak / Block 1
            m_best_pcu, m_best_spd = 0, 0
            for start in m_starts:
                end = start + 60
                tot_s = sum(minute_data.get(m, {}).get(lid, {}).get('S', 0) for m in range(start, end))
                tot_l = sum(minute_data.get(m, {}).get(lid, {}).get('L', 0) for m in range(start, end))
                tot_t = sum(minute_data.get(m, {}).get(lid, {}).get('T', 0) for m in range(start, end))
                tot_sv = sum(minute_data.get(m, {}).get(lid, {}).get('speed_vol', 0) for m in range(start, end))
                tot_v = sum(minute_data.get(m, {}).get(lid, {}).get('vol', 0) for m in range(start, end))

                pcu = tot_s * pce_s + tot_l * pce_l + tot_t * pce_t
                spd = (tot_sv / tot_v) if tot_v > 0 else 0
                if pcu >= m_best_pcu:
                    m_best_pcu = pcu
                    m_best_spd = spd

            # Evening Peak / Block 2 (if present, else use m_starts)
            e_best_pcu, e_best_spd = 0, 0
            for start in (e_starts or m_starts):
                end = start + 60
                tot_s = sum(minute_data.get(m, {}).get(lid, {}).get('S', 0) for m in range(start, end))
                tot_l = sum(minute_data.get(m, {}).get(lid, {}).get('L', 0) for m in range(start, end))
                tot_t = sum(minute_data.get(m, {}).get(lid, {}).get('T', 0) for m in range(start, end))
                tot_sv = sum(minute_data.get(m, {}).get(lid, {}).get('speed_vol', 0) for m in range(start, end))
                tot_v = sum(minute_data.get(m, {}).get(lid, {}).get('vol', 0) for m in range(start, end))

                pcu = tot_s * pce_s + tot_l * pce_l + tot_t * pce_t
                spd = (tot_sv / tot_v) if tot_v > 0 else 0
                if pcu >= e_best_pcu:
                    e_best_pcu = pcu
                    e_best_spd = spd

            final_peak_metrics[lid]['morning'] = {'pcu': m_best_pcu, 'speed': m_best_spd}
            final_peak_metrics[lid]['evening'] = {'pcu': e_best_pcu, 'speed': e_best_spd}
    else:
        # Fullday or custom mode (> 4 hours): Independent 1-hour slots method
        for lid in target_links:
            m_pcu, m_spd = 0, 0
            e_pcu, e_spd = 0, 0
            half = max(1, len(hours_list) // 2)
            for i, h_name in enumerate(hours_list):
                d = hourly_data[h_name][lid]
                pcu = d['S'] * pce_s + d['L'] * pce_l + d['T'] * pce_t
                spd = (d['speed_vol'] / d['vol']) if d['vol'] > 0 else 0
                if i < half:
                    if pcu >= m_pcu:
                        m_pcu, m_spd = pcu, spd
                else:
                    if pcu >= e_pcu:
                        e_pcu, e_spd = pcu, spd

            final_peak_metrics[lid]['morning'] = {'pcu': m_pcu, 'speed': m_spd}
            final_peak_metrics[lid]['evening'] = {'pcu': e_pcu, 'speed': e_spd}
        
    return hourly_data, final_peak_metrics, link_vdid_map

SHOULDER_RANGES = [
    ('國道1號', 'S', 2.79, 5.025),
    ('國道1號', 'S', 18.2, 23.15),
    ('國道1號', 'S', 36.22, 41.0),
    ('國道1號', 'S', 60.88, 62.0),
    ('國道1號', 'S', 62.8, 65.0),
    ('國道1號', 'S', 71.72, 83.0),
    ('國道1號', 'S', 84.5, 86.0),
    ('國道1號', 'S', 87.29, 91.0),
    ('國道1號', 'S', 91.59, 93.032),
    ('國道1號', 'S', 97.45, 99.0),
    ('國道1號', 'S', 107.53, 109.87),
    ('國道1號', 'S', 175.005, 178.015),
    ('國道1號', 'S', 196.765, 198.0),
    ('國道1號', 'S', 300.63, 302.0),
    ('國道1號', 'S', 338.95, 341.003),
    ('國道1號', 'S', 350.25, 354.03),
    ('國道1號', 'S', 36.0, 363.2),
    ('國道1號', 'S', 368.5, 369.0),
    ('國道1號', 'S', 1.49, 2.33),
    ('國道1號', 'S', 2.66, 4.85),
    ('國道1號', 'S', 7.3, 8.84),
    ('國道1號', 'S', 0.5, 11.0),
    ('國道1號', 'S', 17.05, 22.0),
    ('國道1號', 'S', 41.0, 48.08),
    ('國道1號', 'S', 62.0, 64.05),
    ('國道1號', 'N', 87.0, 90.5),
    ('國道1號', 'N', 91.0, 93.01),
    ('國道1號', 'N', 100.0, 109.72),
    ('國道1號', 'N', 161.0, 162.18),
    ('國道1號', 'N', 163.0, 164.6),
    ('國道1號', 'N', 168.0, 172.08),
    ('國道1號', 'N', 320.07, 323.5),
    ('國道1號', 'N', 350.05, 353.2),
    ('國道1號', 'N', 368.0, 368.4),
    ('國道1號', 'N', 367.0, 368.6),
    ('國道1號', 'S', 0.0, 16.266),
    ('國道1號', 'N', 0.0, 20.52),
    ('國道1號', 'N', 20.8, 25.8),
    ('國道1號', 'N', 9.95, 26.41),
    ('國道2號', 'E', 9.68, 10.0),
    ('國道2號', 'E', 15.12, 18.0),
    ('國道2號', 'E', 19.25, 20.0),
    ('國道2號', 'W', 6.04, 8.25),
    ('國道2號', 'W', 11.0, 14.6),
    ('國道3號', 'S', 11.55, 12.0),
    ('國道3號', 'S', 104.0, 109.0),
    ('國道3號', 'S', 381.3, 383.0),
    ('國道3號', 'N', 0.0, 0.71),
    ('國道3號', 'N', 16.55, 17.95),
    ('國道3號', 'N', 43.05, 46.0),
    ('國道3號', 'N', 63.25, 67.65),
    ('國道3號', 'N', 101.55, 103.1),
    ('國道3號', 'N', 16.0, 16.0),
    ('國道3號', 'N', 8.0, 15.99),
    ('國道3號', 'N', 16.0, 16.04),
    ('國道5號', 'S', 4.0, 4.89),
    ('國道10號', 'E', 32.58, 33.72),
]

def check_is_shoulder_open(road_name, road_dir, km):
    road_norm = road_name.replace('國道', '國').replace('號', '').strip()
    if road_norm == '國1':
        road_key = '國道1號'
    elif road_norm == '國3':
        road_key = '國道3號'
    elif road_norm == '國5':
        road_key = '國道5號'
    elif road_norm == '國1高架' or road_norm == '汐五及五楊高架' or road_norm == '國道1號高架道路':
        road_key = '汐五及五楊高架'
    else:
        road_key = road_name
        
    dir_abbr = road_dir
    if road_dir == '北上' or road_dir == 'N':
        dir_abbr = 'N'
    elif road_dir == '南下' or road_dir == 'S':
        dir_abbr = 'S'
    elif road_dir == '往東' or road_dir == 'E':
        dir_abbr = 'E'
    elif road_dir == '往西' or road_dir == 'W':
        dir_abbr = 'W'

    for r_csv, d_dir, min_k, max_k in SHOULDER_RANGES:
        if r_csv == road_key and d_dir == dir_abbr:
            if min_k <= km <= max_k:
                return 'Y'
    return 'N'

def load_interchange_mileages():
    import csv
    import re
    ic_mileage = {}
    
    fw_csv = os.path.join(BASE_DIR, 'reference_files', 'Freeway_Interchanges_Full.csv')
    if os.path.exists(fw_csv):
        try:
            with open(fw_csv, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    cols = list(row.keys())
                    if len(cols) >= 3:
                        road = str(row[cols[0]]).strip().replace('臺', '台')
                        name = str(row[cols[1]]).strip().replace('臺', '台')
                        km_val = str(row[cols[2]]).strip()
                        if road and name and km_val:
                            km_str = re.sub(r'[^0-9.]', '', km_val)
                            if km_str:
                                km = float(km_str)
                                if road not in ic_mileage:
                                    ic_mileage[road] = {}
                                ic_mileage[road][name] = km
                                if '交流道' in name:
                                    ic_mileage[road][name.replace('交流道', '')] = km
                                    ic_mileage[road][name.replace('交流道', '').replace('系統', '')] = km
        except Exception as e:
            print(f"[Warning] Error loading Freeway_Interchanges_Full.csv: {e}")
            
    exp_csv = os.path.join(BASE_DIR, 'reference_files', 'Expressway_Interchanges.csv')
    if os.path.exists(exp_csv):
        try:
            with open(exp_csv, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                header = next(reader)
                for row in reader:
                    if len(row) >= 3:
                        road = str(row[0]).strip().replace('臺', '台')
                        name = str(row[1]).strip().replace('臺', '台')
                        km_val = str(row[2]).strip()
                        if road and name and km_val:
                            km_str = re.sub(r'[^0-9.]', '', km_val)
                            if km_str:
                                km = float(km_str)
                                if road not in ic_mileage:
                                    ic_mileage[road] = {}
                                ic_mileage[road][name] = km
                                if '交流道' in name:
                                    ic_mileage[road][name.replace('交流道', '')] = km
                                    ic_mileage[road][name.replace('交流道', '').replace('系統', '')] = km
        except Exception as e:
            print(f"[Warning] Error loading Expressway_Interchanges.csv: {e}")
            
    for road in ['國道1號', '國道3號', '國道5號', '汐五及五楊高架', '國道1號高架道路']:
        if road not in ic_mileage:
            ic_mileage[road] = {}
    ic_mileage['國道1號']['基隆端'] = 0.0
    ic_mileage['國道1號']['基隆'] = 1.1
    ic_mileage['國道1號']['高雄端'] = 373.0
    ic_mileage['國道1號']['大安溪橋'] = 154.5
    ic_mileage['國道1號']['瑞隆'] = 373.0
    ic_mileage['國道1號']['漁港路'] = 374.0
    ic_mileage['國道3號']['基金'] = 0.0
    ic_mileage['國道3號']['林邊'] = 427.0
    ic_mileage['國道3號']['大鵬灣端'] = 431.5
    ic_mileage['國道5號']['南港系統'] = 0.0
    ic_mileage['國道5號']['蘇澳'] = 54.0
    
    viaduct_ics = {
        '汐止': 13.0, '汐止端': 13.0,
        '堤頂': 18.0,
        '環北': 26.0,
        '下塔悠': 20.0,
        '五股': 31.0, '五股轉接道': 31.0,
        '泰山': 35.8, '泰山轉接道': 35.8,
        '機場系統': 52.5, '機場系統轉接道': 52.5,
        '中壢': 59.0, '中壢轉接道': 59.0,
        '校前路': 70.0,
        '楊梅': 71.0, '楊梅端': 71.0
    }
    ic_mileage['汐五及五楊高架'] = viaduct_ics
    ic_mileage['國道1號高架道路'] = viaduct_ics
    
    return ic_mileage

def resolve_excel_segments(xlsx_path, ic_mileage):
    import re
    resolved_rows = []
    
    if not os.path.exists(xlsx_path):
        print(f"[Warning] Excel file not found: {xlsx_path}")
        return []
        
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        excel_to_csv_road = {
            '國1': '國道1號',
            '國1高架': '國道1號高架道路',
            '國2': '國道2號',
            '國2甲': '國道2號甲線',
            '國3': '國道3號',
            '國3甲': '國道3號甲線',
            '國4': '國道4號',
            '國5': '國道5號',
            '國6': '國道6號',
            '國8': '國道8號',
            '國10': '國道10號',
        }
        
        def resolve_endpoint_km(road_csv, endpoint_desc):
            m = re.search(r'([0-9.]+)[Kk]', endpoint_desc)
            if m:
                return float(m.group(1))
            road_ics = ic_mileage.get(road_csv, {})
            if endpoint_desc in road_ics:
                return road_ics[endpoint_desc]
            sorted_ic_names = sorted(road_ics.keys(), key=lambda x: len(x), reverse=True)
            for ic in sorted_ic_names:
                if ic in endpoint_desc:
                    return road_ics[ic]
            return None

        for r in range(2, ws.max_row + 1):
            road = ws.cell(row=r, column=1).value
            seg = ws.cell(row=r, column=2).value
            if road and seg:
                road_str = str(road).strip().replace('臺', '台')
                seg_str = str(seg).strip().replace('臺', '台')
                road_csv = excel_to_csv_road.get(road_str, road_str)
                
                if seg_str.endswith("區間"):
                    clean_name = seg_str.replace("區間", "").strip()
                    km = resolve_endpoint_km(road_csv, clean_name)
                    if km is not None:
                        resolved_rows.append({
                            'road': road_csv,
                            'segment': seg_str,
                            'min_km': km - 1.3,
                            'max_km': km + 1.3,
                            'is_zone': True,
                            'lanes_south': ws.cell(row=r, column=3).value,
                            'lanes_north': ws.cell(row=r, column=4).value,
                            'speed_south': ws.cell(row=r, column=6).value,
                            'speed_north': ws.cell(row=r, column=7).value
                        })
                        continue
                        
                parts = re.split(r'[～－\-~]', seg_str)
                if len(parts) == 2:
                    km1 = resolve_endpoint_km(road_csv, parts[0].strip())
                    km2 = resolve_endpoint_km(road_csv, parts[1].strip())
                    if km1 is not None and km2 is not None:
                        resolved_rows.append({
                            'road': road_csv,
                            'segment': seg_str,
                            'min_km': min(km1, km2),
                            'max_km': max(km1, km2),
                            'is_zone': False,
                            'lanes_south': ws.cell(row=r, column=3).value,
                            'lanes_north': ws.cell(row=r, column=4).value,
                            'speed_south': ws.cell(row=r, column=6).value,
                            'speed_north': ws.cell(row=r, column=7).value
                        })
                        
    except Exception as e:
        print(f"[Warning] Error parsing 國道車道數、速限.xlsx: {e}")
        
    return resolved_rows

def parse_speed_limit(val, default_limit=100.0):
    if val is None:
        return default_limit
    import re
    val_str = str(val).strip()
    m = re.search(r'最高\s*(\d+)', val_str)
    if m:
        return float(m.group(1))
    m = re.search(r'(\d+)', val_str)
    if m:
        return float(m.group(1))
    return default_limit

def parse_lane_count(val, default_lanes=3):
    if val is None:
        return default_lanes
    val_str = str(val).strip()
    val_str_clean = val_str.replace('車', '').replace('路肩', '').replace(' ', '').strip()
    if '+' in val_str_clean:
        parts = val_str_clean.split('+')
        try:
            return int(float(parts[0])) + int(float(parts[1]))
        except ValueError:
            try:
                return int(float(parts[0]))
            except ValueError:
                return default_lanes
    else:
        try:
            return int(float(val_str_clean))
        except ValueError:
            return default_lanes

def load_link_metadata(vd_xml_path, target_links):
    if not os.path.exists(vd_xml_path):
        return {}
        
    tree = ET.parse(vd_xml_path)
    root = tree.getroot()
    ns = '{http://traffic.transportdata.tw/standard/traffic/schema/}'
    
    metadata = {}
    for vd in root.iter(f'{ns}VD'):
        vd_id = vd.findtext(f'{ns}VDID')
        road_name = vd.findtext(f'{ns}RoadName') or '國道2號'
        location_mile = vd.findtext(f'{ns}LocationMile')
        
        for dlink in vd.iter(f'{ns}DetectionLink'):
            link_id = dlink.findtext(f'{ns}LinkID')
            road_dir = dlink.findtext(f'{ns}RoadDirection')
            lane_num = int(dlink.findtext(f'{ns}LaneNum') or 3)
            
            if link_id in target_links:
                feature_code = link_id[6] if len(link_id) >= 7 else '0'
                is_mainline = (feature_code == '0')
                dir_text = '往東' if road_dir == 'E' else ('往西' if road_dir == 'W' else ('北上' if road_dir == 'N' else '南下'))
                
                metadata[link_id] = {
                    'vd_id': vd_id,
                    'road_name': road_name,
                    'road_dir': dir_text,
                    'lanes': lane_num,
                    'is_mainline': is_mainline,
                    'feature_code': feature_code,
                    'location_mile': location_mile
                }
                
    return metadata

def build_excel_report(output_file, hourly_data, mainline_rows, ramp_rows, target_links, meta, hours_list, link_vdid_map, pce_s=1.0, pce_l=1.8, pce_t=2.5):
    report_progress(85, "Backing up existing files & formatting Excel sheets...")
    backup_file_if_exists(output_file)
    
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border_thin = Border(left=Side(style='thin', color='D9D9D9'),
                         right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'),
                         bottom=Side(style='thin', color='D9D9D9'))
    header_font = Font(name='微軟正黑體', size=11, bold=True)
    data_font = Font(name='微軟正黑體', size=10)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    
    # ------------------ Sheet 1: 國道主線 Analysis ------------------
    ws_main = wb.active
    ws_main.title = '國道主線'
    ws_main.views.sheetView[0].showGridLines = True
    
    ws_main.merge_cells('A1:A2')
    ws_main['A1'] = '道路名稱'
    ws_main.merge_cells('B1:B2')
    ws_main['B1'] = '路段範圍'
    ws_main.merge_cells('C1:C2')
    ws_main['C1'] = '方向'
    ws_main.merge_cells('D1:D2')
    ws_main['D1'] = '道路容量\n(PCPH)'
    ws_main.merge_cells('E1:E2')
    ws_main['E1'] = '速限\n(KPH)'
    
    ws_main.merge_cells('F1:I1')
    ws_main['F1'] = '晨峰'
    ws_main.merge_cells('J1:M1')
    ws_main['J1'] = '昏峰'
    
    sub_headers = ['交通量\n(PCPH)', 'V/C', '速率(KPH)', '服務水準']
    for idx, sh in enumerate(sub_headers):
        ws_main.cell(row=2, column=6+idx, value=sh)
        ws_main.cell(row=2, column=10+idx, value=sh)
        
    for r in range(1, 3):
        for c in range(1, 14):
            cell = ws_main.cell(row=r, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin
            
    row_idx = 3
    for row in mainline_rows:
        ws_main.cell(row=row_idx, column=1, value=row['road_name']).alignment = align_center
        ws_main.cell(row=row_idx, column=2, value=row['segment']).alignment = align_left
        ws_main.cell(row=row_idx, column=3, value=row['direction']).alignment = align_center
        ws_main.cell(row=row_idx, column=4, value=round(row['capacity'])).number_format = '#,##0'
        ws_main.cell(row=row_idx, column=5, value=round(row['speed_limit'])).number_format = '0'
        
        ws_main.cell(row=row_idx, column=6, value=round(row['m_pcu'])).number_format = '#,##0'
        ws_main.cell(row=row_idx, column=7, value=row['m_vc']).number_format = '0.00'
        ws_main.cell(row=row_idx, column=8, value=row['m_speed']).number_format = '0.0'
        ws_main.cell(row=row_idx, column=9, value=row['m_los']).alignment = align_center
        
        ws_main.cell(row=row_idx, column=10, value=round(row['e_pcu'])).number_format = '#,##0'
        ws_main.cell(row=row_idx, column=11, value=row['e_vc']).number_format = '0.00'
        ws_main.cell(row=row_idx, column=12, value=row['e_speed']).number_format = '0.0'
        ws_main.cell(row=row_idx, column=13, value=row['e_los']).alignment = align_center
        # Hidden helper column for link_id (used by web_server for 24hr series matching)
        ws_main.cell(row=row_idx, column=14, value=row['link_id'])
        
        for c in range(1, 14):
            cell = ws_main.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = border_thin
        row_idx += 1
        
    # ------------------ Sheet 2: 國道匝道 Analysis ------------------
    ws_ramp = wb.create_sheet(title='國道匝道')
    ws_ramp.views.sheetView[0].showGridLines = True
    
    ws_ramp.merge_cells('A1:A2')
    ws_ramp['A1'] = '道路名稱'
    ws_ramp.merge_cells('B1:B2')
    ws_ramp['B1'] = '交流道名稱'
    ws_ramp.merge_cells('C1:C2')
    ws_ramp['C1'] = '方向'
    ws_ramp.merge_cells('D1:D2')
    ws_ramp['D1'] = '出入別'
    ws_ramp.merge_cells('E1:E2')
    ws_ramp['E1'] = '往'
    ws_ramp.merge_cells('F1:F2')
    ws_ramp['F1'] = '道路容量\n(PCPH)'
    ws_ramp.merge_cells('G1:G2')
    ws_ramp['G1'] = '速限\n(KPH)'
    
    ws_ramp.merge_cells('H1:K1')
    ws_ramp['H1'] = '晨峰'
    ws_ramp.merge_cells('L1:O1')
    ws_ramp['L1'] = '昏峰'
    
    for idx, sh in enumerate(sub_headers):
        ws_ramp.cell(row=2, column=8+idx, value=sh)
        ws_ramp.cell(row=2, column=12+idx, value=sh)
        
    for r in range(1, 3):
        for c in range(1, 16):
            cell = ws_ramp.cell(row=r, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin
            
    row_idx = 3
    for row in ramp_rows:
        ws_ramp.cell(row=row_idx, column=1, value=row['road_name']).alignment = align_center
        ws_ramp.cell(row=row_idx, column=2, value=row['interchange']).alignment = align_left
        ws_ramp.cell(row=row_idx, column=3, value=row['direction']).alignment = align_center
        ws_ramp.cell(row=row_idx, column=4, value=row['in_out']).alignment = align_center
        ws_ramp.cell(row=row_idx, column=5, value=row['destination']).alignment = align_center
        ws_ramp.cell(row=row_idx, column=6, value=round(row['capacity'])).number_format = '#,##0'
        ws_ramp.cell(row=row_idx, column=7, value=round(row['speed_limit'])).number_format = '0'
        
        ws_ramp.cell(row=row_idx, column=8, value=round(row['m_pcu'])).number_format = '#,##0'
        ws_ramp.cell(row=row_idx, column=9, value=row['m_vc']).number_format = '0.00'
        ws_ramp.cell(row=row_idx, column=10, value=row['m_speed']).number_format = '0.0'
        ws_ramp.cell(row=row_idx, column=11, value=row['m_los']).alignment = align_center
        
        ws_ramp.cell(row=row_idx, column=12, value=round(row['e_pcu'])).number_format = '#,##0'
        ws_ramp.cell(row=row_idx, column=13, value=row['e_vc']).number_format = '0.00'
        ws_ramp.cell(row=row_idx, column=14, value=row['e_speed']).number_format = '0.0'
        ws_ramp.cell(row=row_idx, column=15, value=row['e_los']).alignment = align_center
        # Hidden helper column for link_id (used by web_server for 24hr series matching)
        ws_ramp.cell(row=row_idx, column=16, value=row['link_id'])
        
        for c in range(1, 16):
            cell = ws_ramp.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = border_thin
        row_idx += 1

    # ------------------ Sheets 3 & 4: 24hr Summaries (only for >= 12hr analyses) ------------------
    if len(hours_list) >= 12:
        # Sheet A: 24小時流量與車速對照表
        ws_m = wb.create_sheet(title='24小時流量與車速矩陣')
        ws_m.views.sheetView[0].showGridLines = True

        m_headers = ['LinkID', '道路名稱', '路段/交流道', '方向', '容量(PCPH)', '速限(KPH)']
        for h_name in hours_list:
            m_headers.append(f"{h_name}\n流量(PCU)")
        for h_name in hours_list:
            m_headers.append(f"{h_name}\n速率(KPH)")

        for col_i, h_t in enumerate(m_headers, 1):
            cell = ws_m.cell(row=1, column=col_i, value=h_t)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin

        row_i = 2
        for r in (mainline_rows + ramp_rows):
            lid = r['link_id']
            seg_or_ic = r.get('segment') or r.get('interchange')
            ws_m.cell(row=row_i, column=1, value=lid).alignment = align_center
            ws_m.cell(row=row_i, column=2, value=r['road_name']).alignment = align_center
            ws_m.cell(row=row_i, column=3, value=seg_or_ic).alignment = align_left
            ws_m.cell(row=row_i, column=4, value=r['direction']).alignment = align_center
            ws_m.cell(row=row_i, column=5, value=round(r['capacity'])).number_format = '#,##0'
            ws_m.cell(row=row_i, column=6, value=round(r['speed_limit'])).number_format = '0'

            # Fill PCU columns
            col_offset = 7
            for idx, h_name in enumerate(hours_list):
                d = hourly_data[h_name][lid]
                pcu_val = round(d['S'] * pce_s + d['L'] * pce_l + d['T'] * pce_t)
                c_cell = ws_m.cell(row=row_i, column=col_offset + idx, value=pcu_val)
                c_cell.number_format = '#,##0'

            # Fill Speed columns
            col_offset_spd = col_offset + len(hours_list)
            for idx, h_name in enumerate(hours_list):
                d = hourly_data[h_name][lid]
                spd_val = (d['speed_vol'] / d['vol']) if d['vol'] > 0 else 0
                c_cell = ws_m.cell(row=row_i, column=col_offset_spd + idx, value=spd_val)
                c_cell.number_format = '0.0'

            for c in range(1, len(m_headers) + 1):
                ws_m.cell(row=row_i, column=c).font = data_font
                ws_m.cell(row=row_i, column=c).border = border_thin
            row_i += 1

        # Sheet B: 24小時LOS對照表
        ws_los = wb.create_sheet(title='24小時LOS對照表')
        ws_los.views.sheetView[0].showGridLines = True

        los_headers = ['LinkID', '道路名稱', '路段/交流道', '方向', '容量(PCPH)', '速限(KPH)'] + hours_list
        for col_i, h_t in enumerate(los_headers, 1):
            cell = ws_los.cell(row=1, column=col_i, value=h_t)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin

        los_fills = {
            'A': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
            'B': PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
            'C': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'D': PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
            'E': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            'F': PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"),
        }

        row_i = 2
        for r in (mainline_rows + ramp_rows):
            lid = r['link_id']
            cap = r['capacity']
            limit = r['speed_limit']
            seg_or_ic = r.get('segment') or r.get('interchange')

            ws_los.cell(row=row_i, column=1, value=lid).alignment = align_center
            ws_los.cell(row=row_i, column=2, value=r['road_name']).alignment = align_center
            ws_los.cell(row=row_i, column=3, value=seg_or_ic).alignment = align_left
            ws_los.cell(row=row_i, column=4, value=r['direction']).alignment = align_center
            ws_los.cell(row=row_i, column=5, value=round(cap)).number_format = '#,##0'
            ws_los.cell(row=row_i, column=6, value=round(limit)).number_format = '0'

            for idx, h_name in enumerate(hours_list):
                d = hourly_data[h_name][lid]
                pcu_val = d['S'] * pce_s + d['L'] * pce_l + d['T'] * pce_t
                spd_val = (d['speed_vol'] / d['vol']) if d['vol'] > 0 else 0
                vc_val = pcu_val / cap if cap > 0 else 0
                los_code = calculate_los(vc_val, spd_val, limit)

                l_cell = ws_los.cell(row=row_i, column=7 + idx, value=los_code)
                l_cell.alignment = align_center
                letter = los_code[0] if los_code else 'A'
                if letter in los_fills:
                    l_cell.fill = los_fills[letter]

            for c in range(1, len(los_headers) + 1):
                ws_los.cell(row=row_i, column=c).font = data_font
                ws_los.cell(row=row_i, column=c).border = border_thin
            row_i += 1

    # ------------------ Sheets 5+: Raw Hourly Data ------------------
    raw_headers = [
        '偵測器代碼 (VDID)', '路段代碼 (LinkID)', 
        '小客車數量 (S)', '小客車速率 (KPH)',
        '大客車數量 (L)', '大客車速率 (KPH)',
        '聯結車數量 (T)', '聯結車速率 (KPH)',
        '總車輛數 (車)', '總當量 (PCU)', '加權速率 (KPH)'
    ]
    
    for h_name in hours_list:
        ws = wb.create_sheet(title=h_name)
        ws.views.sheetView[0].showGridLines = True
        
        for col_idx, h_text in enumerate(raw_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin
            
        row_idx = 2
        for lid in target_links:
            vd_info = meta.get(lid, {})
            vd_id = link_vdid_map.get(lid) or vd_info.get('vd_id', '')
            
            d = hourly_data[h_name][lid]
            s_vol = round(d['S'])
            s_spd = (d['S_speed_vol'] / d['S']) if d['S'] > 0 else 0
            
            l_vol = round(d['L'])
            l_spd = (d['L_speed_vol'] / d['L']) if d['L'] > 0 else 0
            
            t_vol = round(d['T'])
            t_spd = (d['T_speed_vol'] / d['T']) if d['T'] > 0 else 0
            
            tot_veh = round(d['S'] + d['L'] + d['T'])
            tot_pcu = round(d['S'] * pce_s + d['L'] * pce_l + d['T'] * pce_t)
            avg_spd = (d['speed_vol'] / d['vol']) if d['vol'] > 0 else 0
            
            ws.cell(row=row_idx, column=1, value=vd_id).alignment = align_center
            ws.cell(row=row_idx, column=2, value=lid).alignment = align_center
            ws.cell(row=row_idx, column=3, value=s_vol).number_format = '#,##0'
            ws.cell(row=row_idx, column=4, value=s_spd).number_format = '0.0'
            ws.cell(row=row_idx, column=5, value=l_vol).number_format = '#,##0'
            ws.cell(row=row_idx, column=6, value=l_spd).number_format = '0.0'
            ws.cell(row=row_idx, column=7, value=t_vol).number_format = '#,##0'
            ws.cell(row=row_idx, column=8, value=t_spd).number_format = '0.0'
            ws.cell(row=row_idx, column=9, value=tot_veh).number_format = '#,##0'
            ws.cell(row=row_idx, column=10, value=tot_pcu).number_format = '#,##0'
            ws.cell(row=row_idx, column=11, value=avg_spd).number_format = '0.0'
            
            for c in range(1, 12):
                ws.cell(row=row_idx, column=c).font = data_font
                ws.cell(row=row_idx, column=c).border = border_thin
            row_idx += 1

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
    report_progress(95, "Saving Excel workbook to disk...")
    wb.save(output_file)
    print(f"Report successfully saved to {output_file}")
    report_progress(100, "Execution completed successfully!")


def main(date_str='20260716', mode='peak', custom_hours_str='', pce_s=1.0, pce_l=1.8, pce_t=2.5):
    timestamp_log = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"run_{date_str}_{timestamp_log}.log"
    log_filepath = os.path.join(LOGS_DIR, log_filename)
    sys.stdout = DualLogger(log_filepath)
    
    report_progress(5, f"Starting Traffic Analysis for Date: {date_str} (Mode: {mode}, PCE: S={pce_s}/L={pce_l}/T={pce_t})")
    print(f"Runtime Log File: {log_filepath}")
    
    dt = datetime.datetime.strptime(date_str, "%Y%m%d")
    is_weekend = dt.weekday() in [5, 6]
    
    if mode == 'fullday':
        print("Date type: Full Day (00:00 - 24:00, 24 Hours)")
        hours_list = [f"{h:02d}00-{(h+1):02d}00" for h in range(24)]
        period_map = {h: f"{h:02d}00-{(h+1):02d}00" for h in range(24)}
        download_hours = list(range(24))
    elif mode == 'custom' and custom_hours_str:
        # Supports "7,8,17,18" or range syntax like "7-9, 17-19"
        blocks = parse_custom_blocks(custom_hours_str)
        chours = sorted(set(h for b in blocks for h in b))
        if not chours:
            chours = [7, 8, 17, 18]
        hours_list = [f"{h:02d}00-{(h+1):02d}00" for h in chours]
        period_map = {h: f"{h:02d}00-{(h+1):02d}00" for h in chours}
        download_hours = chours
        print(f"Date type: Custom Hours ({chours})")
    else:
        if is_weekend:
            print("Date type: Weekend / Holiday Peak (10-12 Morning, 16-18 Evening)")
            hours_list = ['1000-1100', '1100-1200', '1600-1700', '1700-1800']
            period_map = {10: '1000-1100', 11: '1100-1200', 16: '1600-1700', 17: '1700-1800'}
            download_hours = [10, 11, 16, 17]
        else:
            print("Date type: Weekday Peak (07-09 Morning, 17-19 Evening)")
            hours_list = ['0700-0800', '0800-0900', '1700-1800', '1800-1900']
            period_map = {7: '0700-0800', 8: '0800-0900', 17: '1700-1800', 18: '1800-1900'}
            download_hours = [7, 8, 17, 18]
        
    report_progress(10, "Loading target link IDs configuration...")
    with open(TARGET_LINKS_FILE, 'r', encoding='utf-8') as f:
        target_links = [line.strip() for line in f if line.strip()]
        
    print(f"Loaded {len(target_links)} target LinkIDs.")
    
    date_dir = download_vd_data(date_str, download_hours)
    hourly_data, metrics, link_vdid_map = process_vd_data(date_dir, target_links, (hours_list, period_map), is_weekend, mode, custom_hours_str, pce_s, pce_l, pce_t)
    meta = load_link_metadata(VD_POINT_LIST_FILE, target_links)
    
    mainline_rows = []
    ramp_rows = []
    
    ic_mileage = load_interchange_mileages()
    resolved_segments = resolve_excel_segments(os.path.join(BASE_DIR, 'reference_files', '國道車道數、速限.xlsx'), ic_mileage)
    
    for lid in target_links:
        m_data = metrics.get(lid, {'morning': {'pcu': 0, 'speed': 0}, 'evening': {'pcu': 0, 'speed': 0}})
        info = meta.get(lid)
        
        if not info:
            prefix = lid[:5] if len(lid) >= 5 else ''
            road_map = {
                '00001': '國道1號',
                '00002': '國道2號',
                '00003': '國道3號',
                '00004': '國道4號',
                '00005': '國道5號',
                '00006': '國道6號',
                '00008': '國道8號',
                '00010': '國道10號',
            }
            road_name = road_map.get(prefix, '國道2號')
            feature_code = lid[6] if len(lid) >= 7 else '0'
            is_mainline = (feature_code == '0')
            direction_char = lid[7] if len(lid) >= 8 else '0'
            
            is_east_west = any(r in road_name for r in ('2號', '4號', '6號', '8號', '10號', '2甲', '3甲'))
            if is_east_west:
                road_dir = '往東' if direction_char == '0' else '往西'
            else:
                road_dir = '南下' if direction_char == '0' else '北上'
            
            info = {
                'vd_id': f"VD-{lid}",
                'road_name': road_name,
                'road_dir': road_dir,
                'lanes': 3,
                'is_mainline': is_mainline,
                'feature_code': feature_code,
                'location_mile': None
            }
            
        m_pcu = m_data['morning']['pcu']
        m_spd = m_data['morning']['speed']
        e_pcu = m_data['evening']['pcu']
        e_spd = m_data['evening']['speed']
        
        link_km = None
        location_mile = info.get('location_mile')
        if location_mile:
            import re
            m = re.match(r'(\d+)K\+(\d+)', str(location_mile).strip())
            if m:
                link_km = float(m.group(1)) + float(m.group(2)) / 1000.0
                
        if link_km is None:
            if len(lid) >= 13:
                try:
                    mile_str = lid[8:13]
                    if info['is_mainline']:
                        link_km = float(mile_str) / 100.0
                    else:
                        link_km = float(mile_str) / 1000.0
                except ValueError:
                    link_km = 0.0
            else:
                link_km = 0.0
                
        if info['is_mainline']:
            matched_seg = None
            road_norm = info['road_name'].replace('臺', '台').strip()
            if road_norm == '國1高架':
                road_norm = '國道1號高架道路'
                
            for r_seg in resolved_segments:
                if r_seg['road'] == road_norm and r_seg['is_zone']:
                    if r_seg['min_km'] <= link_km <= r_seg['max_km']:
                        matched_seg = r_seg
                        break
            if not matched_seg:
                for r_seg in resolved_segments:
                    if r_seg['road'] == road_norm and not r_seg['is_zone']:
                        if r_seg['min_km'] <= link_km <= r_seg['max_km']:
                            matched_seg = r_seg
                            break
                            
            if matched_seg:
                seg_name = matched_seg['segment'].replace('區間', '').strip()
                is_north_west = info['road_dir'] in ('北上', '往西', 'N', 'W')
                raw_lanes = matched_seg['lanes_north'] if is_north_west else matched_seg['lanes_south']
                raw_speed = matched_seg['speed_north'] if is_north_west else matched_seg['speed_south']
                
                default_spd = 100.0
                if '3號' in road_norm or '國3' in road_norm:
                    default_spd = 110.0
                elif '5號' in road_norm or '國5' in road_norm or '8號' in road_norm or '國8' in road_norm or '10號' in road_norm or '國10' in road_norm:
                    default_spd = 80.0 if link_km <= 15.0 else 90.0
                elif '國1高架' in road_norm or '五楊' in road_norm or '汐五' in road_norm:
                    default_spd = 100.0
                    
                limit = parse_speed_limit(raw_speed, default_spd)
                lanes = parse_lane_count(raw_lanes, info['lanes'])
            else:
                seg_name = f"{info['road_name']} {info['road_dir']} {link_km:.1f}K"
                lanes = info['lanes']
                
                default_spd = 100.0
                if '3號' in info['road_name'] or '國3' in info['road_name']:
                    default_spd = 110.0
                elif '5號' in info['road_name'] or '國5' in info['road_name'] or '8號' in info['road_name'] or '國8' in info['road_name'] or '10號' in info['road_name'] or '國10' in info['road_name']:
                    default_spd = 80.0 if link_km <= 15.0 else 90.0
                limit = default_spd
                
            open_shoulder = check_is_shoulder_open(info['road_name'], info['road_dir'], link_km)
            cap = get_mainline_capacity(limit, lanes, open_shoulder)
            
            m_vc = m_pcu / cap if cap > 0 else 0
            e_vc = e_pcu / cap if cap > 0 else 0
            
            mainline_rows.append({
                'link_id': lid,
                'road_name': info['road_name'],
                'segment': seg_name,
                'direction': info['road_dir'],
                'capacity': cap,
                'speed_limit': limit,
                'm_pcu': m_pcu,
                'm_vc': m_vc,
                'm_speed': m_spd,
                'm_los': calculate_los(m_vc, m_spd, limit),
                'e_pcu': e_pcu,
                'e_vc': e_vc,
                'e_speed': e_spd,
                'e_los': calculate_los(e_vc, e_spd, limit)
            })
        else:
            road_ics = ic_mileage.get(info['road_name'], {})
            closest_ic = "交流道"
            min_dist = 9999.9
            for ic_name, ic_km in road_ics.items():
                dist = abs(ic_km - link_km)
                if dist < min_dist:
                    min_dist = dist
                    closest_ic = ic_name
                    
            if not closest_ic.endswith('交流道') and not closest_ic.endswith('系統') and not closest_ic.endswith('端'):
                closest_ic = closest_ic + '交流道'
                
            vd_id_upper = str(info['vd_id']).upper()
            if '-I-' in vd_id_upper or '匝-入' in vd_id_upper or '匝-加速' in vd_id_upper or '入' in vd_id_upper:
                in_out = '入口'
            elif '-O-' in vd_id_upper or '匝-出' in vd_id_upper or '匝-減速' in vd_id_upper or '出' in vd_id_upper:
                in_out = '出口'
            else:
                in_out = '出口'
                
            if lid == '0000201101040H':
                in_out = '入口'
            elif lid == '0000201001000H':
                in_out = '出口'
                
            if "環" in str(info['vd_id']) or "LOOP" in vd_id_upper:
                limit = 40.0
            else:
                limit = 50.0
                
            lanes = info['lanes']
            cap = get_ramp_capacity(in_out, lanes)
            dest = closest_ic.replace('交流道', '').strip()
            
            m_vc = m_pcu / cap if cap > 0 else 0
            e_vc = e_pcu / cap if cap > 0 else 0
            
            ramp_rows.append({
                'link_id': lid,
                'road_name': info['road_name'],
                'interchange': closest_ic,
                'direction': info['road_dir'],
                'in_out': in_out,
                'destination': dest,
                'capacity': cap,
                'speed_limit': limit,
                'm_pcu': m_pcu,
                'm_vc': m_vc,
                'm_speed': m_spd,
                'm_los': calculate_los(m_vc, m_spd, limit),
                'e_pcu': e_pcu,
                'e_vc': e_vc,
                'e_speed': e_spd,
                'e_los': calculate_los(e_vc, e_spd, limit)
            })

    timestamp_hhmm = datetime.datetime.now().strftime("%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"VD_traffic_report_{date_str}{timestamp_hhmm}.xlsx")
    build_excel_report(output_path, hourly_data, mainline_rows, ramp_rows, target_links, meta, hours_list, link_vdid_map, pce_s, pce_l, pce_t)

if __name__ == '__main__':
    date_input = sys.argv[1] if len(sys.argv) > 1 else '20260716'
    mode_input = sys.argv[2] if len(sys.argv) > 2 else 'peak'
    custom_hours_input = sys.argv[3] if len(sys.argv) > 3 else ''
    pce_s_input = float(sys.argv[4]) if len(sys.argv) > 4 else 1.0
    pce_l_input = float(sys.argv[5]) if len(sys.argv) > 5 else 1.8
    pce_t_input = float(sys.argv[6]) if len(sys.argv) > 6 else 2.5
    main(date_input, mode_input, custom_hours_input, pce_s_input, pce_l_input, pce_t_input)

