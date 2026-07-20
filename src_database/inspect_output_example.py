import openpyxl
import pandas as pd
import os

file_path = 'OUTPUT表格範例.xlsx'

def inspect_example():
    if not os.path.exists(file_path):
        print(f"File {file_path} does not exist.")
        return
        
    print(f"=== Inspecting {file_path} ===")
    wb = openpyxl.load_workbook(file_path)
    print("Sheets:", wb.sheetnames)
    
    output = []
    output.append(f"Workbook: {file_path}")
    output.append(f"Sheets: {wb.sheetnames}")
    
    xl = pd.ExcelFile(file_path)
    for sname in xl.sheet_names:
        df = xl.parse(sname)
        output.append(f"\n--- Sheet: '{sname}' (Dimensions: {df.shape}) ---")
        output.append("Columns: " + str(df.columns.tolist()))
        output.append("First 15 rows:")
        output.append(df.head(15).to_string())
        
    scratch_dir = r"C:\Users\Owner\.gemini\antigravity\brain\5da4f6b9-54ae-48d5-bc02-be8a5a3438dd\scratch"
    out_txt = os.path.join(scratch_dir, "output_example_inspection.txt")
    with open(out_txt, 'w', encoding='utf-8') as f:
        f.write("\n".join(output))
    print("Wrote inspection to", out_txt)

if __name__ == '__main__':
    inspect_example()
